import json
import mimetypes
import os
import shutil
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Any

import docx
import pandas as pd
import pypdf
from PIL import ExifTags, Image

from file_pilot.analysis.archive_reader import read_archive_index
from file_pilot.analysis.image_describer import describe_image, format_image_description_result

DEFAULT_MAX_LEN = 300
DEFAULT_LIST_DEPTH = 1
DEFAULT_LIST_CHAR_LIMIT = 1800
DIR_INSPECT_DEPTH = 2
DIR_INSPECT_CHAR_LIMIT = 800
LIST_TRUNCATION_NOTICE = "...[目录结果过长已截断]"
TEXT_ENCODINGS = ["utf-8", "utf-8-sig", "gbk", "utf-16"]
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
CSV_EXTENSIONS = {".csv"}
AUDIO_EXTENSIONS = {
    ".aac",
    ".flac",
    ".m4a",
    ".mp3",
    ".ogg",
    ".opus",
    ".wav",
    ".wma",
}
VIDEO_EXTENSIONS = {".avi", ".mkv", ".mov", ".mp4", ".webm"}
MEDIA_BINARY_EXTENSIONS = AUDIO_EXTENSIONS | VIDEO_EXTENSIONS
COMMON_BINARY_EXTENSIONS = {
    ".7z",
    ".bin",
    ".cbr",
    ".cbz",
    ".dll",
    ".dmg",
    ".exe",
    ".iso",
    ".jar",
    ".msi",
    ".rar",
    ".tar",
    ".tgz",
    ".webp",
}
KNOWN_BINARY_EXTENSIONS = MEDIA_BINARY_EXTENSIONS | COMMON_BINARY_EXTENSIONS | IMAGE_EXTENSIONS


def _normalize_local_path(path: str) -> str:
    return os.path.normpath(path or ".")


def _is_allowed_local_path(path: str, allowed_base_dir: str | None = None) -> bool:
    if not allowed_base_dir:
        return True

    resolved_path = os.path.abspath(path)
    resolved_base = os.path.abspath(allowed_base_dir)
    try:
        common_path = os.path.commonpath([resolved_path, resolved_base])
    except ValueError:
        return False
    return common_path == resolved_base


def _read_text_with_fallback(filepath: str) -> str:
    last_error = None
    for encoding in TEXT_ENCODINGS:
        try:
            with open(filepath, "r", encoding=encoding) as file:
                return file.read()
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise last_error
    raise UnicodeDecodeError("utf-8", b"", 0, 1, "unable to decode file")


def _file_basic_info(filepath: str) -> dict[str, str | int]:
    stat = os.stat(filepath)
    mime_type = mimetypes.guess_type(filepath)[0] or "unknown"
    return {
        "size_bytes": stat.st_size,
        "modified_time": datetime.fromtimestamp(stat.st_mtime).replace(microsecond=0).isoformat(),
        "extension": os.path.splitext(filepath)[1].lower() or "无扩展名",
        "mime_type": mime_type,
    }


def _format_basic_info(filepath: str, *, unsupported_reason: str | None = None) -> str:
    info = _file_basic_info(filepath)
    lines = [
        "系统信息：",
        f"- 文件大小：{info['size_bytes']} bytes",
        f"- 修改时间：{info['modified_time']}",
        f"- 扩展名：{info['extension']}",
        f"- MIME 类型：{info['mime_type']}",
    ]
    if unsupported_reason:
        lines.append(f"- 无法直读原因：{unsupported_reason}")
    return "\n".join(lines)


def _metadata_value(metadata: Any, keys: list[str]) -> str:
    for key in keys:
        try:
            value = metadata.get(key)
        except Exception:
            value = None
        if value is None:
            continue
        if isinstance(value, list):
            value = ", ".join(str(item) for item in value if str(item).strip())
        text = str(value).strip()
        if text:
            return text
    return ""


def read_audio_metadata(filepath: str) -> str:
    """提取音频文件的轻量标签和媒体属性。"""
    lines = [_format_basic_info(filepath)]
    try:
        import mutagen  # type: ignore[import-not-found]
    except Exception:
        lines.append("音频元数据：未安装 mutagen，当前仅返回基础文件属性。")
        return "\n".join(lines)

    try:
        audio = mutagen.File(filepath, easy=True)
        if audio is None:
            lines.append("音频元数据：无法识别该音频格式或标签为空。")
            return "\n".join(lines)

        fields = [
            ("标题", ["title"]),
            ("艺术家", ["artist"]),
            ("专辑", ["album"]),
            ("专辑艺术家", ["albumartist", "album_artist"]),
            ("曲目号", ["tracknumber"]),
            ("碟号", ["discnumber"]),
            ("年份", ["date", "year"]),
            ("类型", ["genre"]),
        ]
        lines.append("音频元数据：")
        extracted = False
        for label, keys in fields:
            value = _metadata_value(audio, keys)
            if not value:
                continue
            lines.append(f"- {label}: {value}")
            extracted = True

        info = getattr(audio, "info", None)
        if info is not None:
            duration = getattr(info, "length", None)
            bitrate = getattr(info, "bitrate", None)
            sample_rate = getattr(info, "sample_rate", None)
            channels = getattr(info, "channels", None)
            if duration is not None:
                lines.append(f"- 时长: {round(float(duration), 2)} 秒")
                extracted = True
            if bitrate is not None:
                lines.append(f"- 比特率: {int(bitrate)} bps")
                extracted = True
            if sample_rate is not None:
                lines.append(f"- 采样率: {int(sample_rate)} Hz")
                extracted = True
            if channels is not None:
                lines.append(f"- 声道数: {int(channels)}")
                extracted = True

        if not extracted:
            lines.append("- 未读取到可用于整理的常见音频标签。")
        return "\n".join(lines)
    except Exception as exc:
        lines.append(f"音频元数据：读取失败: {exc}")
        return "\n".join(lines)


def read_video_metadata(filepath: str) -> str:
    """通过 ffprobe 提取视频文件的轻量媒体属性。"""
    lines = [_format_basic_info(filepath)]
    ffprobe = shutil.which("ffprobe")
    if not ffprobe:
        lines.append("视频元数据：未找到 ffprobe，当前仅返回基础文件属性。")
        return "\n".join(lines)

    try:
        completed = subprocess.run(
            [
                ffprobe,
                "-v",
                "error",
                "-print_format",
                "json",
                "-show_format",
                "-show_streams",
                filepath,
            ],
            check=True,
            capture_output=True,
            text=True,
            timeout=5,
        )
        payload = json.loads(completed.stdout or "{}")
        format_info = payload.get("format") if isinstance(payload, dict) else {}
        streams = payload.get("streams") if isinstance(payload, dict) else []
        video_stream = next(
            (
                stream
                for stream in streams or []
                if isinstance(stream, dict) and stream.get("codec_type") == "video"
            ),
            {},
        )
        audio_stream = next(
            (
                stream
                for stream in streams or []
                if isinstance(stream, dict) and stream.get("codec_type") == "audio"
            ),
            {},
        )

        lines.append("视频元数据：")
        duration = str(format_info.get("duration") or video_stream.get("duration") or "").strip()
        bitrate = str(format_info.get("bit_rate") or video_stream.get("bit_rate") or "").strip()
        if duration:
            try:
                duration = str(round(float(duration), 2))
            except ValueError:
                pass
            lines.append(f"- 时长: {duration} 秒")
        if bitrate:
            lines.append(f"- 比特率: {bitrate} bps")
        if video_stream:
            codec = str(video_stream.get("codec_name") or "").strip()
            width = video_stream.get("width")
            height = video_stream.get("height")
            if codec:
                lines.append(f"- 视频编码: {codec}")
            if width and height:
                lines.append(f"- 分辨率: {width}x{height}")
        if audio_stream:
            codec = str(audio_stream.get("codec_name") or "").strip()
            channels = audio_stream.get("channels")
            sample_rate = str(audio_stream.get("sample_rate") or "").strip()
            if codec:
                lines.append(f"- 音频编码: {codec}")
            if channels:
                lines.append(f"- 音频声道数: {channels}")
            if sample_rate:
                lines.append(f"- 音频采样率: {sample_rate} Hz")
        if len(lines) <= 2:
            lines.append("- 未读取到可用于整理的常见视频属性。")
        return "\n".join(lines)
    except Exception as exc:
        lines.append(f"视频元数据：读取失败: {exc}")
        return "\n".join(lines)


def _truncate_text_head_tail(text: str, max_len: int) -> str:
    if max_len <= 0 or len(text) <= max_len:
        return text
    notice = "\n...[中间内容已省略]...\n"
    available = max_len - len(notice)
    if available <= 0:
        return text[:max_len]
    head_len = max(1, available // 2)
    tail_len = max(1, available - head_len)
    return text[:head_len].rstrip() + notice + text[-tail_len:].lstrip()


def _join_limited_lines(lines: list[str], char_limit: int, truncation_notice: str = LIST_TRUNCATION_NOTICE) -> str:
    full_text = "\n".join(lines)
    if char_limit <= 0 or len(full_text) <= char_limit:
        return full_text

    suffix = f"\n{truncation_notice}"
    available = char_limit - len(suffix)
    if available <= 0:
        return truncation_notice[:char_limit]

    kept_lines: list[str] = []
    current_length = 0
    for line in lines:
        extra = len(line) if not kept_lines else len(line) + 1
        if current_length + extra > available:
            break
        kept_lines.append(line)
        current_length += extra

    if not kept_lines:
        return full_text[:available].rstrip() + suffix
    return "\n".join(kept_lines) + suffix


def read_pdf(filepath, max_len=DEFAULT_MAX_LEN):
    """提取 PDF 文本内容。"""
    try:
        reader = pypdf.PdfReader(filepath)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
            if len(text) >= max_len:
                break
        return text.strip()
    except Exception as exc:
        return f"读取 PDF 失败: {exc}"


def read_docx(filepath, max_len=DEFAULT_MAX_LEN):
    """提取 Word 文本内容。"""
    try:
        document = docx.Document(filepath)
        text = ""
        for para in document.paragraphs:
            text += para.text + "\n"
            if len(text) >= max_len:
                break
        return text.strip()
    except Exception as exc:
        return f"读取 Word 失败: {exc}"


def read_excel(filepath, max_len=DEFAULT_MAX_LEN):
    """提取 Excel 内容摘要。"""
    try:
        output = []
        total_len = 0
        if os.path.splitext(filepath)[1].lower() in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook

                workbook_meta = load_workbook(filepath, read_only=True, data_only=True)
                try:
                    for worksheet in workbook_meta.worksheets:
                        rows = list(worksheet.iter_rows(min_row=1, max_row=11, values_only=True))
                        headers = [
                            str(value).strip()
                            for value in (rows[0] if rows else [])
                            if value is not None and str(value).strip()
                        ]
                        sample_rows = rows[1:] if rows else []
                        dataframe = pd.DataFrame(sample_rows, columns=headers or None)
                        combined = (
                            f"Sheet: {worksheet.title}\n"
                            f"总行数: {max(0, int(worksheet.max_row or 0) - 1)}\n"
                            f"总列数: {int(worksheet.max_column or 0)}\n"
                            f"列名: {', '.join(headers) if headers else '无'}\n"
                            f"前 10 行样例:\n{dataframe.to_string(index=False)}\n"
                        )
                        output.append(combined)
                        total_len += len(combined)
                        if total_len >= max_len:
                            break
                finally:
                    workbook_meta.close()
                return "".join(output)
            except Exception as exc:
                return f"读取 Excel 失败: {exc}"

        with pd.ExcelFile(filepath) as workbook:
            for sheet_name in workbook.sheet_names:
                dataframe = pd.read_excel(workbook, sheet_name=sheet_name, nrows=10)
                column_names = [str(column) for column in dataframe.columns]
                combined = (
                    f"Sheet: {sheet_name}\n"
                    f"样例行数: {len(dataframe)}\n"
                    f"样例列数: {len(dataframe.columns)}\n"
                    f"列名: {', '.join(column_names) if column_names else '无'}\n"
                    f"前 10 行样例:\n{dataframe.to_string(index=False)}\n"
                )
                output.append(combined)
                total_len += len(combined)
                if total_len >= max_len:
                    break

        return "".join(output)
    except Exception as exc:
        return f"读取 Excel 失败: {exc}"


def read_csv(filepath, max_len=DEFAULT_MAX_LEN):
    """提取 CSV 内容摘要。"""
    try:
        sample = pd.read_csv(filepath, nrows=10)
        column_names = [str(column) for column in sample.columns]
        total_rows = 0
        for chunk in pd.read_csv(filepath, chunksize=10000):
            total_rows += len(chunk)
        content = (
            f"CSV 表格\n"
            f"总行数: {total_rows}\n"
            f"总列数: {len(column_names)}\n"
            f"列名: {', '.join(column_names) if column_names else '无'}\n"
            f"前 10 行样例:\n{sample.to_string(index=False)}"
        )
        return _truncate_text_head_tail(content, max_len)
    except Exception as exc:
        return f"读取 CSV 失败: {exc}"


def read_image_metadata(filepath: str) -> str:
    """提取图片基础元数据和常见 EXIF 信息。"""
    try:
        with Image.open(filepath) as image:
            lines = [
                "图片系统信息：",
                f"- 格式：{image.format or 'unknown'}",
                f"- 尺寸：{image.width}x{image.height}",
                f"- 色彩模式：{image.mode}",
            ]
            exif = image.getexif()
            if exif:
                tag_names = {
                    "DateTimeOriginal",
                    "DateTime",
                    "Make",
                    "Model",
                    "LensModel",
                    "Software",
                    "Orientation",
                    "GPSInfo",
                }
                readable_exif = []
                for tag_id, value in exif.items():
                    tag_name = ExifTags.TAGS.get(tag_id, str(tag_id))
                    if tag_name not in tag_names:
                        continue
                    if isinstance(value, bytes):
                        value = value[:80].hex()
                    readable_exif.append(f"- {tag_name}: {value}")
                if readable_exif:
                    lines.append("EXIF：")
                    lines.extend(readable_exif[:12])
                else:
                    lines.append("EXIF：未提取到常用整理线索")
            else:
                lines.append("EXIF：无")
            return "\n".join(lines)
    except Exception as exc:
        return f"读取图片元数据失败: {exc}"


def list_local_files(directory=".", max_depth=DEFAULT_LIST_DEPTH, char_limit=DEFAULT_LIST_CHAR_LIMIT):
    """列出指定目录下一层内的目录和文件摘要。"""
    try:
        directory = _normalize_local_path(directory)
        if not _is_allowed_local_path(directory):
            return "错误：基于安全考虑，仅允许查看当前目录或 test 子目录下的内容。"
        if not os.path.exists(directory):
            return f"错误：目录 {directory} 不存在。"
        if not os.path.isdir(directory):
            return f"错误：{directory} 不是目录。"

        lines = ["路径 | 类型 | 说明"]
        root_depth = directory.count(os.sep)

        top_level_entries = sorted(
            (entry for entry in os.scandir(directory) if not entry.name.startswith(".")),
            key=lambda entry: entry.name.lower(),
        )
        lines.append(f"{directory} | dir | 包含 {len(top_level_entries)} 个条目")

        for entry in top_level_entries:
            relative_path = os.path.join(directory, entry.name).replace("\\", "/")
            if entry.is_dir():
                child_entries = sorted(
                    (child for child in os.scandir(entry.path) if not child.name.startswith(".")),
                    key=lambda child: child.name.lower(),
                )
                lines.append(f"{relative_path} | dir | 包含 {len(child_entries)} 个条目")

                if max_depth >= 1:
                    current_depth = entry.path.count(os.sep) - root_depth
                    if current_depth <= max_depth:
                        for child in child_entries:
                            child_path = os.path.join(relative_path, child.name).replace("\\", "/")
                            if child.is_dir():
                                lines.append(f"{child_path} | dir | 已达到递归深度限制")
                            else:
                                suffix = os.path.splitext(child.name)[1].lower() or "无扩展名"
                                lines.append(f"{child_path} | file | {suffix}")
            else:
                suffix = os.path.splitext(entry.name)[1].lower() or "无扩展名"
                lines.append(f"{relative_path} | file | {suffix}")

        return _join_limited_lines(lines, char_limit=char_limit)
    except Exception as exc:
        return f"无法列出目录 {directory}: {exc}"


def read_local_file(filename, max_len=DEFAULT_MAX_LEN, allowed_base_dir: str | None = None):
    """读取本地文件内容。"""
    try:
        filename = _normalize_local_path(filename)
        if not _is_allowed_local_path(filename, allowed_base_dir=allowed_base_dir):
            return "错误：基于安全考虑，本程序仅限读取允许目录内的文件。"
        if not os.path.exists(filename):
            return f"错误：文件 {filename} 不存在。"
        if os.path.isdir(filename):
            structure = list_local_files(filename, max_depth=DIR_INSPECT_DEPTH, char_limit=DIR_INSPECT_CHAR_LIMIT)
            return f"--- 目录 [{filename}] 结构 ---\n{structure}\n--- 结构结束 ---"

        ext = os.path.splitext(filename)[1].lower()
        if ext == ".pdf":
            content = read_pdf(filename, max_len=max_len)
        elif ext in [".docx", ".doc"]:
            content = read_docx(filename, max_len=max_len)
        elif ext in [".xlsx", ".xls"]:
            content = read_excel(filename, max_len=max_len)
        elif ext in CSV_EXTENSIONS:
            content = read_csv(filename, max_len=max_len)
        elif ext == ".zip":
            content = read_archive_index(filename, max_entries=max_len)
        elif ext in IMAGE_EXTENSIONS:
            content = read_image_metadata(filename) + "\n" + format_image_description_result(describe_image(filename))
        elif ext in AUDIO_EXTENSIONS:
            content = read_audio_metadata(filename)
        elif ext in VIDEO_EXTENSIONS:
            content = read_video_metadata(filename)
        elif ext in KNOWN_BINARY_EXTENSIONS:
            content = _format_basic_info(
                filename,
                unsupported_reason="该文件是已知二进制格式，当前扫描工具不会按文本解码读取其内容。",
            )
        else:
            content = _read_text_with_fallback(filename)

        if len(content) > max_len:
            content = _truncate_text_head_tail(content, max_len) + "\n...[内容过长已截断]"

        return f"--- 文件 [{filename}] 内容开始 ---\n{content}\n--- 内容结束 ---"
    except UnicodeDecodeError:
        return "该文件可能是二进制格式或使用了非 UTF-8 编码，请检查文件后缀是否正确。"
    except Exception as exc:
        return f"无法读取文件 {filename}: {exc}"


BATCH_READ_SEPARATOR = "\n\n"


def read_local_files_batch(
    filenames: list[str],
    max_len: int = DEFAULT_MAX_LEN,
    allowed_base_dir: str | None = None,
) -> str:
    """批量探查多个条目的内容摘要或目录结构，减少多次工具调用开销。"""
    if not filenames:
        return "错误：未提供任何文件名。"

    worker_count = max(1, min(4, len(filenames)))
    results: list[str] = [""] * len(filenames)
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(read_local_file, filename, max_len=max_len, allowed_base_dir=allowed_base_dir): index
            for index, filename in enumerate(filenames)
        }
        for future in as_completed(futures):
            index = futures[future]
            try:
                results[index] = future.result()
            except Exception as exc:  # pragma: no cover - defensive fallback
                results[index] = f"无法读取文件 {filenames[index]}: {exc}"

    return BATCH_READ_SEPARATOR.join(results)

